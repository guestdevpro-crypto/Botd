import telebot
from telebot import types
from datetime import datetime, date
import os
from openpyxl import Workbook, load_workbook

# ================= CONFIG =================
TOKEN = "8349098130:AAGFzgtd9JgbSOHxBhvYC03HV82A_nUNDkY"

# Bir nechta admin ID
ADMINS = [1732314424, 8355611778]

SINF_RAHBARI_ID = 1732314424
MAKTAB_NAZORATCHI_ID = 8355611778

EXCEL_FILE = "davomat.xlsx"
USERS_FILE = "users.txt"

VALID_CLASSES = ["8-A","8-B","9-A","9-B","10-A","10-B","11-A"]

bot = telebot.TeleBot(TOKEN)
user_data = {}

# ================= INIT =================
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Sana","Vaqt","UserID","Ism","Telefon","Sinf","Sabab"])
        wb.save(EXCEL_FILE)

def save_user(user_id):
    if not os.path.exists(USERS_FILE):
        open(USERS_FILE,"w").close()
    with open(USERS_FILE,"r+") as f:
        users = f.read().splitlines()
        if str(user_id) not in users:
            f.write(str(user_id)+"\n")

def save_to_excel(data, user_id):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        now = datetime.now()
        ws.append([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            user_id,
            data.get('name','Noma\'lum'),
            data.get('phone','Noma\'lum'),
            data.get('class','Noma\'lum'),
            data.get('reason','Noma\'lum')
        ])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Excel saqlashda xatolik: {e}")

def has_sent_today(user_id):
    try:
        if not os.path.exists(EXCEL_FILE):
            return False
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        today = date.today().strftime("%Y-%m-%d")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == today and row[2] == user_id:
                return True
        return False
    except:
        return False

# ================= USER FLOW =================
@bot.message_handler(commands=['start'])
def start(message):
    save_user(message.chat.id)
    if has_sent_today(message.chat.id):
        bot.send_message(message.chat.id,"❌ Siz bugun yuborgansiz.")
        return
    user_data[message.chat.id] = {}
    bot.send_message(message.chat.id,"📝 Ism familiyangizni yozing:")

@bot.message_handler(func=lambda m: m.chat.id in user_data and 'name' not in user_data[m.chat.id])
def get_name(message):
    user_data[message.chat.id]['name'] = message.text
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(types.KeyboardButton("📱 Telefon yuborish", request_contact=True))
    bot.send_message(message.chat.id,"📞 Telefonni yuboring:", reply_markup=kb)

@bot.message_handler(content_types=['contact'])
def get_phone(message):
    try:
        user_data[message.chat.id]['phone'] = message.contact.phone_number
    except:
        bot.send_message(message.chat.id,"❌ Iltimos, telefon kontaktini yuboring.")
        return
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for cls in VALID_CLASSES:
        kb.add(types.KeyboardButton(cls))
    bot.send_message(message.chat.id,"🏫 Sinfni tanlang:", reply_markup=kb)

@bot.message_handler(func=lambda m: m.text in VALID_CLASSES)
def get_class(message):
    user_data[message.chat.id]['class'] = message.text
    bot.send_message(message.chat.id,"✍️ Sabab (kamida 7 so‘z):", reply_markup=types.ReplyKeyboardRemove())

@bot.message_handler(func=lambda m: m.chat.id in user_data and 'class' in user_data[m.chat.id] and 'reason' not in user_data[m.chat.id])
def get_reason(message):
    if len(message.text.split()) < 7:
        bot.send_message(message.chat.id,"❌ Kamida 7 so‘z yozing.")
        return
    user_data[message.chat.id]['reason'] = message.text
    kb = types.InlineKeyboardMarkup()
    kb.row(
        types.InlineKeyboardButton("👨‍🏫 Rahbar", callback_data="teacher"),
        types.InlineKeyboardButton("🏫 Nazoratchi", callback_data="school")
    )
    bot.send_message(message.chat.id,"Yuborilsinmi?", reply_markup=kb)

@bot.callback_query_handler(func=lambda call: call.data in ["teacher","school"])
def send_data(call):
    data = user_data.get(call.message.chat.id)
    if not data: return
    text = (
        f"📌 Kelmagan\n"
        f"👤 {data.get('name','Noma\'lum')}\n"
        f"📞 {data.get('phone','Noma\'lum')}\n"
        f"🏫 {data.get('class','Noma\'lum')}\n"
        f"📝 {data.get('reason','Noma\'lum')}"
    )
    try:
        save_to_excel(data, call.message.chat.id)
    except:
        pass
    if call.data == "teacher":
        bot.send_message(SINF_RAHBARI_ID, text)
    else:
        bot.send_message(MAKTAB_NAZORATCHI_ID, text)
    bot.edit_message_text("✅ Yuborildi!", call.message.chat.id, call.message.message_id)

# ================= ADMIN PANEL =================
@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if message.chat.id not in ADMINS:
        bot.send_message(message.chat.id,"❌ Siz admin emassiz.")
        return
    kb = types.InlineKeyboardMarkup()
    kb.row(
        types.InlineKeyboardButton("📊 Jami", callback_data="stats"),
        types.InlineKeyboardButton("📅 Bugun", callback_data="today")
    )
    kb.row(
        types.InlineKeyboardButton("🏫 Sinf statistikasi", callback_data="class_stats"),
        types.InlineKeyboardButton("👥 Foydalanuvchilar", callback_data="users")
    )
    kb.row(
        types.InlineKeyboardButton("📈 Grafik", callback_data="graph"),
        types.InlineKeyboardButton("📥 Excel", callback_data="excel")
    )
    bot.send_message(message.chat.id,"🔐 Admin Panel", reply_markup=kb)

@bot.callback_query_handler(func=lambda call: call.data in ["stats","today","class_stats","users","graph","excel"])
def admin_actions(call):
    try:
        if call.message.chat.id not in ADMINS:
            return
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        if call.data == "stats":
            total = ws.max_row - 1
            bot.send_message(call.message.chat.id, f"📊 Jami: {total} ta")
        elif call.data == "today":
            today = datetime.now().strftime("%Y-%m-%d")
            count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == today)
            bot.send_message(call.message.chat.id, f"📅 Bugun: {count} ta")
        elif call.data == "class_stats":
            stats = {}
            for r in ws.iter_rows(min_row=2, values_only=True):
                cls = r[5] if r[5] in VALID_CLASSES else "Noma'lum"
                stats[cls] = stats.get(cls,0)+1
            text = "🏫 Sinflar bo‘yicha davomat:\n\n"
            for cls,count in stats.items():
                text += f"{cls}: {count} ta\n"
            bot.send_message(call.message.chat.id,text)
        elif call.data == "users":
            if os.path.exists(USERS_FILE):
                with open(USERS_FILE) as f:
                    users = f.read()
                bot.send_message(call.message.chat.id,"👥 Foydalanuvchilar:\n"+users)
        elif call.data == "graph":
            stats = {}
            for r in ws.iter_rows(min_row=2, values_only=True):
                cls = r[5] if r[5] in VALID_CLASSES else "Noma'lum"
                stats[cls] = stats.get(cls,0)+1
            text = "📊 Davomat Grafikasi (sinf bo‘yicha)\n\n"
            for cls,count in stats.items():
                bar = "█"*count
                text += f"{cls:5} | {bar} {count} ta\n"
            bot.send_message(call.message.chat.id,text)
        elif call.data == "excel":
            bot.send_document(call.message.chat.id, open(EXCEL_FILE,"rb"))
    except:
        bot.send_message(call.message.chat.id,"❌ Xatolik yuz berdi, keyinroq urinib ko‘ring.")

# ================= RUN =================
init_excel()
print("Bot ishlayapti...")
bot.infinity_polling()
from flask import Flask
import threading
import os

app = Flask(__name__)

@app.route('/')
def home():
    return "Bot ishlayapti"

def run_flask():
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))

# Flaskni alohida oqimda ishga tushiramiz
threading.Thread(target=run_flask).start()

# Bot polling
bot.infinity_polling()